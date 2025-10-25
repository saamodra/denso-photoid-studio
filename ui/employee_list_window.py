import sys
import os
import datetime
from PyQt6.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, 
                             QTableWidgetItem, QPushButton, QLineEdit, QLabel, QComboBox, 
                             QFileDialog, QMessageBox, QHeaderView, QAbstractItemView, QDialog)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QFont, QColor
from modules.database import db_manager
from modules.session_manager import session_manager

from ui.employee_detail_window import EmployeeDetailPage
from ui.dialogs.admin_respon_request_dialog import AdminResponseDialog

# Pustaka untuk Excel, pastikan sudah diinstal: pip install openpyxl
import openpyxl

class CustomDialog(QDialog):
    """Custom dialog with consistent styling"""
    def __init__(self, parent=None, title="", message="", buttons=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setFixedSize(400, 200)

        # Main layout
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Message label
        self.message_label = QLabel(message)
        self.message_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.message_label.setWordWrap(True)
        layout.addWidget(self.message_label)

        # Button layout
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        if buttons is None:
            buttons = [("OK", QDialog.DialogCode.Accepted)]

        self.buttons = []
        for text, role in buttons:
            btn = QPushButton(text)
            btn.clicked.connect(lambda checked, r=role: self.done(r))
            button_layout.addWidget(btn)
            self.buttons.append(btn)

        layout.addLayout(button_layout)

        # Apply styling
        self.setStyleSheet("""
            QDialog {
                background-color: #FFFFFF;
                color: #333333;
            }
            QLabel {
                background-color: #FFFFFF;
                color: #333333;
                font-size: 14px;
                padding: 10px;
            }
            QPushButton {
                background-color: #E60012;
                color: #FFFFFF;
                font-weight: bold;
                font-size: 14px;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #CC0010;
            }
            QPushButton:pressed {
                background-color: #99000C;
            }
            QPushButton#cancelButton {
                background-color: #6c757d;
            }
            QPushButton#cancelButton:hover {
                background-color: #5a6268;
            }
            QPushButton#cancelButton:pressed {
                background-color: #495057;
            }
        """)


class EmployeeListPage(QWidget):
    """
    Halaman untuk menampilkan, mencari, dan mengelola daftar karyawan
    dengan fitur paginasi, impor, dan ekspor.
    """

    back_button_click = pyqtSignal()  # Emit user data on successful login

    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Data & State Management
        self.all_employees = []  # Data mentah dari database
        self.displayed_employees = []  # Data setelah difilter/dicari
        self.current_page = 1
        self.rows_per_page = 20

        self.init_ui()
        self.apply_styles()

        self.employee_detail_window = None
        # self.load_data() # Ganti dengan koneksi database Anda
        self.current_user = session_manager.get_current_user()

    def showEvent(self, event):
        super().showEvent(event)
        print("Reloading data on show...")
        self.load_data()

    def init_ui(self):
        """Inisialisasi semua komponen UI."""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # --- Header: Tombol Kembali, Judul, Aksi Global ---
        header_layout = QHBoxLayout()
        
        self.back_button = QPushButton("← Kembali")
        self.back_button.setObjectName("BackButton")
        # TODO: Hubungkan sinyal 'clicked' ke fungsi navigasi Anda
        self.back_button.clicked.connect(self.handle_back_button)
        header_layout.addWidget(self.back_button)

        header_layout.addStretch()

        title_label = QLabel("Manajemen Data Karyawan")
        title_label.setObjectName("TitleLabel")
        header_layout.addWidget(title_label)

        header_layout.addStretch()

        self.import_button = QPushButton("📥 Impor Data")
        self.import_button.clicked.connect(self.import_from_excel)
        header_layout.addWidget(self.import_button)

        self.export_button = QPushButton("📤 Ekspor Data")
        self.export_button.clicked.connect(self.export_to_excel)
        header_layout.addWidget(self.export_button)

        # --- Search Bar ---
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Cari berdasarkan NPK atau Nama...")
        self.search_input.textChanged.connect(self.perform_search)
        search_layout.addWidget(self.search_input)

        # --- Table Widget ---
        self.table = QTableWidget()
        # self.table.setColumnCount(12) # 10 dari DB + Status + Aksi
        self.table.setColumnCount(10) # 10 dari DB + Status + Aksi
        self.table.setHorizontalHeaderLabels([
            "No",
            "NPK", 
            "Nama", 
            # "Role", 
            "Seksi", 
            "Departemen", 
            "Perusahaan", 
            "Plant", 
            # "Akses Terakhir", 
            "Foto Terakhir",
            "Status Request", 
            # "File Foto",
            "Aksi"
        ])
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

        # khusus kolom aksi
        self.table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)

        # self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive) # Nama
        self.table.verticalHeader().setDefaultSectionSize(50)


        # --- Pagination Controls ---
        pagination_layout = QHBoxLayout()
        pagination_layout.setAlignment(Qt.AlignmentFlag.AlignRight)
        
        self.rows_per_page_combo = QComboBox()
        self.rows_per_page_combo.addItems(["10", "20", "50", "100"])
        self.rows_per_page_combo.setCurrentText(str(self.rows_per_page))
        self.rows_per_page_combo.currentIndexChanged.connect(self.change_rows_per_page)
        pagination_layout.addWidget(QLabel("Baris per halaman:"))
        pagination_layout.addWidget(self.rows_per_page_combo)
        pagination_layout.addSpacing(20)

        self.prev_button = QPushButton("< Sebelumnya")
        self.prev_button.clicked.connect(self.go_to_previous_page)
        pagination_layout.addWidget(self.prev_button)

        self.page_label = QLabel("Halaman 1 dari 1")
        self.page_label.setObjectName("PageLabel")
        pagination_layout.addWidget(self.page_label)

        self.next_button = QPushButton("Berikutnya >")
        self.next_button.clicked.connect(self.go_to_next_page)
        pagination_layout.addWidget(self.next_button)

        # --- Add all layouts to main layout ---
        main_layout.addLayout(header_layout)
        main_layout.addLayout(search_layout)
        main_layout.addWidget(self.table)
        main_layout.addLayout(pagination_layout)

    def handle_back_button(self):
        self.back_button_click.emit()

    def apply_styles(self):
        """Menerapkan styling ke seluruh widget."""
        primary_color = "#E60012"
        dark_font_color = "#212121"
        light_font_color = "#757575"
        background_color = "#FFFFFF"
        border_color = "#E0E0E0"

        self.setStyleSheet(f"""
            QWidget {{
                background-color: {background_color};
                color: {dark_font_color};
                font-family: 'Segoe UI';
                font-size: 14px;
            }}
            #TitleLabel {{
                font-size: 24px;
                font-weight: bold;
            }}
            QLineEdit {{
                border: 1px solid {border_color};
                border-radius: 5px;
                padding: 8px;
                background-color: #F9F9F9;
            }}
            QLineEdit:focus {{
                border: 1px solid {primary_color};
            }}
            QPushButton {{
                background-color: {primary_color};
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 5px;
                padding: 10px 15px;
            }}
            QPushButton:hover {{
                background-color: #C30010;
            }}
            QPushButton:pressed {{
                background-color: #A9000E;
            }}
            #BackButton {{
                background-color: {border_color};
                color: {dark_font_color};
            }}
            #BackButton:hover {{
                background-color: #BDBDBD;
            }}
            QTableWidget {{
                border: 1px solid {border_color};
                gridline-color: {border_color};
            }}
            QHeaderView::section {{
                background-color: {primary_color};
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
            }}
            QTableWidget::item {{
                padding: 8px;
            }}
            QTableWidget::item:selected {{
                background-color: #FEE7E9;
                color: {dark_font_color};
            }}
            #PageLabel, QComboBox, QLabel {{
                font-size: 14px;
            }}
            QComboBox {{
                border: 1px solid {border_color};
                padding: 5px;
                border-radius: 5px;
            }}
        """)
    
    def load_data(self):
        """Memuat data dummy. Ganti fungsi ini dengan koneksi database Anda."""
        # TODO: Ganti bagian ini dengan query ke database Anda
        self.all_employees = db_manager.get_all_users_with_request_histories()
        
        # for i in range(1, 101):
        #     self.all_employees.append({
        #         "npk": f"00{i:03}", "name": f"Karyawan {i}", "password": "pass",
        #         "role": "Operator" if i % 2 == 0 else "Staff", "section_id": "SEC01",
        #         "section_name": "Produksi A", "department_id": "DEP01", "department_name": "Manufaktur",
        #         "company": "PT. ABC", "plant": "Cikarang",
        #         "last_access": datetime.datetime.now(), "last_take_photo": datetime.datetime.now(),
        #         "photo_filename": f"photo_{i}.jpg", "card_filename": f"card_{i}.png"
        #     })
        self.displayed_employees = self.all_employees
        self.current_page = 1
        self.refresh_table()

    def refresh_table(self):
        """Memperbarui tampilan tabel berdasarkan halaman dan data saat ini."""
        self.table.setRowCount(0)
        
        total_rows = len(self.displayed_employees)
        total_pages = (total_rows + self.rows_per_page - 1) // self.rows_per_page
        if total_pages == 0: total_pages = 1
        
        self.page_label.setText(f"Halaman {self.current_page} dari {total_pages}")
        
        start_index = (self.current_page - 1) * self.rows_per_page
        end_index = start_index + self.rows_per_page
        
        page_data = self.displayed_employees[start_index:end_index]
        
        self.table.setRowCount(len(page_data))
        
        for row_index, employee in enumerate(page_data):
            # Kolom dari database
            self.table.setItem(row_index, 0, QTableWidgetItem(str(row_index + 1)))
            self.table.setItem(row_index, 1, QTableWidgetItem(str(employee.get("npk", ""))))
            self.table.setItem(row_index, 2, QTableWidgetItem(str(employee.get("name", ""))))
            # self.table.setItem(row_index, 2, QTableWidgetItem(str(employee.get("role", ""))))
            self.table.setItem(row_index, 3, QTableWidgetItem(str(employee.get("section_name", ""))))
            self.table.setItem(row_index, 4, QTableWidgetItem(str(employee.get("department_name", ""))))
            self.table.setItem(row_index, 5, QTableWidgetItem(str(employee.get("company", ""))))
            self.table.setItem(row_index, 6, QTableWidgetItem(str(employee.get("plant", ""))))
            # self.table.setItem(row_index, 7, QTableWidgetItem(str(employee.get("last_access", ""))))
            self.table.setItem(row_index, 7, QTableWidgetItem(str(employee.get("last_take_photo", ""))))
            # self.table.setItem(row_index, 10, QTableWidgetItem(employee.get("photo_filename", "")))
            
            # Kolom Status (logika custom)
            approved_button = False
            status_teks = str(employee.get("status_request", ""))

            match status_teks:
                case 'requested':
                    status_warna_teks = "black"
                    status_warna_bg = "#FFF176"  # kuning muda
                    approved_button = True
                case 'approved':
                    status_warna_teks = "white"
                    status_warna_bg = "#66BB6A"  # hijau lembut
                case 'rejected':
                    status_warna_teks = "white"
                    status_warna_bg = "#EF5350"  # merah lembut
                case _:
                    status_warna_teks = "white"
                    status_warna_bg = "#9E9E9E"  # abu-abu

            status_item = QTableWidgetItem(status_teks)
            status_item.setForeground(QColor(status_warna_teks))  # warna teks
            status_item.setBackground(QColor(status_warna_bg))    # warna latar
            
            self.table.setItem(row_index, 8,  status_item)

            # Kolom Aksi dengan tiga tombol
            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(0, 0, 0, 0)
            action_layout.setSpacing(6)

            # Tombol Detail
            detail_button = QPushButton("Detail 📄")
            detail_button.setToolTip("Lihat Detail")
            detail_button.clicked.connect(lambda ch, employee=employee: self.show_employee_detail(employee))

            # Tombol Hapus
            delete_button = QPushButton("Hapus 🗑️")
            delete_button.setToolTip("Hapus Data")
            delete_button.clicked.connect(lambda ch, npk=employee.get("npk"): self.delete_employee(npk))

            # Tombol Konfirmasi
            confirmation_button = QPushButton("Setujui ✅")
            confirmation_button.setToolTip("Beri Respon pada Request")
            confirmation_button.clicked.connect(lambda ch, employee=employee: self.confirm_employee(employee, self.current_user))
            confirmation_button.setEnabled(approved_button)  # hanya aktif jika status 'requested'

            # Tambahkan tombol ke layout
            action_layout.addWidget(detail_button)
            action_layout.addWidget(delete_button)
            action_layout.addWidget(confirmation_button)

            # Masukkan widget ke kolom aksi
            self.table.setCellWidget(row_index, 9, action_widget)


            self.table.setCellWidget(row_index, 9, action_widget)

            
        self.prev_button.setEnabled(self.current_page > 1)
        self.next_button.setEnabled(self.current_page < total_pages)

    # --- Pagination Methods ---
    def go_to_next_page(self):
        total_rows = len(self.displayed_employees)
        total_pages = (total_rows + self.rows_per_page - 1) // self.rows_per_page
        if self.current_page < total_pages:
            self.current_page += 1
            self.refresh_table()

    def go_to_previous_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.refresh_table()

    def change_rows_per_page(self):
        self.rows_per_page = int(self.rows_per_page_combo.currentText())
        self.current_page = 1
        self.refresh_table()

    # --- Search Method ---
    def perform_search(self):
        search_text = self.search_input.text().lower()
        if not search_text:
            self.displayed_employees = self.all_employees
        else:
            self.displayed_employees = [
                emp for emp in self.all_employees
                if search_text in emp.get("npk", "").lower() or \
                   search_text in emp.get("name", "").lower()
            ]
        self.current_page = 1
        self.refresh_table()

    # --- Import/Export Methods ---
    def export_to_excel(self):
        if not self.all_employees:
            self.show_styled_messagebox("Informasi", "Tidak ada data untuk diekspor.")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Simpan File Excel", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            headers = list(self.all_employees[0].keys())

            # Remove the password field if it exists
            # if "password" in headers:
            #     headers.remove("password")

            sheet.append(headers)
            
            for employee in self.all_employees:
                row = [employee.get(h, "") for h in headers]
                sheet.append(row)
            
            workbook.save(file_path)
            self.show_styled_messagebox("Sukses", f"Data berhasil diekspor ke {file_path}")
        except Exception as e:
            self.show_styled_messagebox("Error", f"Gagal mengekspor data: {e}", icon=QMessageBox.Icon.Critical)

    def import_from_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Pilih File Excel", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            headers = [cell.value for cell in sheet[1]]
            
            new_employees = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                employee_data = dict(zip(headers, row))
                new_employees.append(employee_data)
                

            if not new_employees:
                self.show_styled_messagebox("Informasi", "Tidak ada data baru ditemukan di dalam file.")
                return

            reply = self.show_styled_messagebox(
                "Konfirmasi Impor",
                f"Anda akan mengimpor {len(new_employees)} data karyawan baru. Lanjutkan?",
                buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                # TODO: Lakukan validasi data sebelum memasukkan ke database
                # TODO: Masukkan data baru ini ke database Anda
                for user in new_employees:
                    db_manager.create_user(user)

                self.all_employees.extend(new_employees)
                self.perform_search() # Untuk me-refresh tampilan
                self.show_styled_messagebox("Sukses", f"{len(new_employees)} data berhasil diimpor.")
        
        except Exception as e:
            self.show_styled_messagebox("Error", f"Gagal mengimpor data: {e}", icon=QMessageBox.Icon.Critical)
    
    # --- Show detail ---
    def show_employee_detail(self, employee):
        dialog = EmployeeDetailPage(employee)
        dialog.exec()

    # --- Confirm employee ---
    def confirm_employee(self, request, admin):
        # self.show_styled_messagebox("Information", f"confirm_employee {npk} is clicked")
        dialog = AdminResponseDialog(request_data=request, admin_user=admin)
        dialog.request_updated.connect(self.load_data) # refresh data
        dialog.exec()


    # --- Confirm employee ---
    def delete_employee(self, npk):
        confirm_dialog = CustomDialog(
            self,
            "Konfirmasi",
            f"Apakah anda yakin menghapus data {npk}?",
            [("Tidak", QDialog.DialogCode.Rejected), ("Ya, hapus", QDialog.DialogCode.Accepted)]
        )

        # Set cancel button styling
        if len(confirm_dialog.buttons) >= 1:
            confirm_dialog.buttons[0].setObjectName("cancelButton")  # "Tidak" button
        if len(confirm_dialog.buttons) >= 2:
            confirm_dialog.buttons[1].setObjectName("confirmButton")  # "Ya" button

        result = confirm_dialog.exec()
        if result == QDialog.DialogCode.Accepted:
            db_manager.remove_user(npk)
            
            self.all_employees = db_manager.get_all_users_with_request_histories()
            self.perform_search() # Untuk me-refresh tampilan
        

    def show_styled_messagebox(self, title, text, icon=QMessageBox.Icon.Information, buttons=QMessageBox.StandardButton.Ok):
        """Menampilkan QMessageBox dengan styling yang konsisten."""
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle(title)
        msg_box.setText(text)
        msg_box.setIcon(icon)
        msg_box.setStandardButtons(buttons)
        
        # Terapkan styling
        msg_box.setStyleSheet(f"""
            QMessageBox {{
                background-color: #FFFFFF;
            }}
            QLabel {{
                color: #212121;
                font-size: 14px;
            }}
            QPushButton {{
                background-color: #E60012;
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                min-width: 80px;
            }}
        """)
        return msg_box.exec()


# --- Untuk Menjalankan Jendela Ini Secara Mandiri (Testing) ---
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = EmployeeListPage()
    # window.showFullScreen()  # tampilkan dalam mode full screen
    # window.show()
    sys.exit(app.exec())